import streamlit as st
import pdfplumber
import re
import io
import tempfile
import os
import datetime
from openpyxl import load_workbook

st.set_page_config(page_title="GSTR-3B → Excel Mapper", page_icon="📊", layout="wide")

# ─────────────────────────────────────────────────────────────────────────────
# NUMBER CLEANING
# ─────────────────────────────────────────────────────────────────────────────

def cn(val):
    """
    Parse a PDF cell value to float. Handles:
    - Leading watermark letters: 'F\n4754151.\n00' → 4754151.00
    - Broken decimals: '4754151.\n00' → 4754151.00
    - Line-wrapped large numbers: '21156221\n0.00' → 211562210.00
    - Dash/empty → None
    """
    if val is None:
        return None
    s = str(val).strip().replace(",", "").replace("₹", "")
    # Normalize all whitespace to single space
    s = " ".join(s.split())
    # Strip leading single-char PDF watermark letters (E, F, I, D …)
    s = re.sub(r'^[A-Z]\s+', '', s).strip()
    # Remove ALL remaining spaces (handles line-wrapped numbers like '21156221 0.00')
    s = s.replace(" ", "")
    if s in ("", "-", "--", "N/A"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


# ─────────────────────────────────────────────────────────────────────────────
# EXTRACTION
# ─────────────────────────────────────────────────────────────────────────────

def parse_itc_table(rows):
    """Parse Section 4 ITC table rows → dict keyed by 4A1..4D2."""
    itc = {}
    label_map = [
        ("4A1",  "(1) import of goods"),
        ("4A2",  "(2) import of services"),
        ("4A3",  "(3) inward supplies liable to reverse charge (other"),
        ("4A4",  "(4) inward supplies from isd"),
        ("4A5",  "(5) all other itc"),
        ("4B1",  "(1) as per rules 38"),
        ("4B2",  "(2) others"),
        ("4C",   "c. net itc available"),
        ("4D",   "(d) other details"),
        ("4D1",  "(1) itc reclaimed"),
        ("4D2",  "(2) ineligible itc"),
    ]
    for row in rows:
        if not row:
            continue
        label = str(row[0] or "").lower().strip()
        nums = [cn(c) for c in row[1:5]]        # IGST, CGST, SGST, CESS
        while len(nums) < 4:
            nums.append(None)
        for key, pattern in label_map:
            if pattern in label:
                itc[key] = {"igst": nums[0], "cgst": nums[1], "sgst": nums[2], "cess": nums[3]}
                break
    return itc


def parse_payment_table(rows):
    """
    Parse Section 6.1 payment table.
    PDF columns (0-indexed): [0]=desc [1]=payable [2]=adj [3]=net
      [4]=ITC_IGST [5]=ITC_CGST [6]=ITC_SGST [7]=ITC_CESS [8]=cash [9]=interest [10]=latefee
    For 6B rows (RCM), ITC columns are dashes; [8]=cash paid.
    """
    pay = {"6A": {}, "6B": {}}
    section = None

    for row in rows:
        if not row:
            continue
        row_text = " ".join(str(c or "") for c in row).lower()
        if "(a) other than reverse charge" in row_text:
            section = "6A"
            continue
        if "(b) reverse charge" in row_text:
            section = "6B"
            continue
        if section is None:
            continue

        first = str(row[0] or "").lower().replace("\n", " ").strip()
        if "integrated" in first:
            tax = "igst"
        elif "central" in first:
            tax = "cgst"
        elif "state" in first or "ut" in first:
            tax = "sgst"
        elif "cess" in first:
            tax = "cess"
        else:
            continue

        vals = [cn(c) for c in row]
        while len(vals) < 11:
            vals.append(None)

        if section == "6A":
            pay["6A"][tax] = {
                "payable":  vals[1],
                "itc_igst": vals[4],
                "itc_cgst": vals[5],
                "itc_sgst": vals[6],
                "itc_cess": vals[7],
                "cash":     vals[8],
                "interest": vals[9],
                "latefee":  vals[10],
            }
        else:  # 6B
            pay["6B"][tax] = {
                "payable": vals[1],
                "cash":    vals[8],
            }

    return pay


def extract_gstr3b(pdf_bytes):
    """Extract all GSTR-3B sections using table-based parsing."""
    meta = {"gstin": None, "period": None, "year": None,
             "legal_name": None, "arn": None, "date_arn": None}
    all_tables = []

    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            for tbl in (page.extract_tables() or []):
                all_tables.append(tbl)

    # ── Header (tables 1 & 2 on page 1) ──
    for tbl in all_tables[:4]:
        for row in tbl:
            if not row or len(row) < 2:
                continue
            key = str(row[0] or "").strip()
            val = str(row[1] or "").strip() if len(row) > 1 else ""
            if key == "Year":
                meta["year"] = val
            elif key == "Period":
                meta["period"] = val
            elif "GSTIN of the supplier" in key:
                m = re.search(r'\b(\d{2}[A-Z]{5}\d{4}[A-Z]\d[Z][A-Z\d])\b', val)
                if m:
                    meta["gstin"] = m.group(1)
            elif "Legal name" in key:
                meta["legal_name"] = val
            elif "ARN" in key and "Date" not in key:
                meta["arn"] = val
            elif "Date of ARN" in key:
                meta["date_arn"] = val

    # ── Section 3.1 ──
    sec31 = {}
    for tbl in all_tables:
        for row in tbl:
            if not row:
                continue
            label = str(row[0] or "").lower()
            vals = [cn(c) for c in row[1:6]]
            while len(vals) < 5:
                vals.append(None)
            if "(a) outward taxable supplies (other than zero" in label:
                sec31["a"] = {"taxable": vals[0], "igst": vals[1], "cgst": vals[2], "sgst": vals[3], "cess": vals[4]}
            elif "(b) outward taxable supplies (zero rated)" in label:
                sec31["b"] = {"taxable": vals[0], "igst": vals[1], "cess": vals[4]}
            elif "(c" in label and "nil rated" in label:
                sec31["c"] = {"taxable": vals[0]}
            elif "(d) inward supplies (liable to reverse charge)" in label:
                sec31["d"] = {"taxable": vals[0], "igst": vals[1], "cgst": vals[2], "sgst": vals[3], "cess": vals[4]}
            elif "(e) non-gst outward" in label:
                sec31["e"] = {"taxable": vals[0]}

    # ── Section 3.1.1 (9(5) ecomm) ──
    sec311 = {}
    for tbl in all_tables:
        for row in tbl:
            if not row:
                continue
            label = str(row[0] or "").lower()
            vals = [cn(c) for c in row[1:6]]
            while len(vals) < 5:
                vals.append(None)
            if "electronic commerce operator pays tax" in label:
                sec311["i"] = {"taxable": vals[0], "igst": vals[1], "cgst": vals[2], "sgst": vals[3], "cess": vals[4]}
            elif "registered person through electronic" in label:
                sec311["ii"] = {"taxable": vals[0]}

    # ── Section 3.2 ──
    sec32 = {}
    for tbl in all_tables:
        for row in tbl:
            if not row or len(row) < 3:
                continue
            label = str(row[0] or "").lower()
            if "unregistered persons" in label:
                sec32["urd"] = {"taxable": cn(row[1]), "igst": cn(row[2])}
            elif "composition taxable" in label:
                sec32["comp"] = {"taxable": cn(row[1]), "igst": cn(row[2])}
            elif "uin holders" in label:
                sec32["uin"] = {"taxable": cn(row[1]), "igst": cn(row[2])}

    # ── Section 4 ITC (split across page 1→2 boundary) ──
    itc_rows = []
    collecting = False
    for tbl in all_tables:
        for row in tbl:
            if not row:
                continue
            label = str(row[0] or "").lower()
            if not collecting and ("itc available" in label or "(1) import of goods" in label
                                   or "(4) inward supplies from isd" in label):
                collecting = True
            if collecting:
                itc_rows.append(row)
                if "(2) ineligible itc" in label:
                    collecting = False
    itc = parse_itc_table(itc_rows)

    # ── Section 5 ──
    sec5 = {}
    for tbl in all_tables:
        for row in tbl:
            if not row:
                continue
            label = str(row[0] or "").lower()
            if "composition scheme" in label and "exempt" in label:
                sec5["comp_inter"]  = cn(row[1]) if len(row) > 1 else None
                sec5["comp_intra"]  = cn(row[2]) if len(row) > 2 else None
            elif "non gst supply" in label:
                sec5["nongst_inter"] = cn(row[1]) if len(row) > 1 else None
                sec5["nongst_intra"] = cn(row[2]) if len(row) > 2 else None

    # ── Section 5.1 Interest & Late Fee ──
    sec51 = {}
    for tbl in all_tables:
        for row in tbl:
            if not row:
                continue
            label = str(row[0] or "").lower()
            if "interest paid" in label:
                sec51["int_igst"] = cn(row[1]) if len(row) > 1 else None
                sec51["int_cgst"] = cn(row[2]) if len(row) > 2 else None
                sec51["int_sgst"] = cn(row[3]) if len(row) > 3 else None
                sec51["int_cess"] = cn(row[4]) if len(row) > 4 else None
            elif "late fee" in label:
                sec51["lf_igst"]  = cn(row[1]) if len(row) > 1 else None
                sec51["lf_cgst"]  = cn(row[2]) if len(row) > 2 else None
                sec51["lf_sgst"]  = cn(row[3]) if len(row) > 3 else None
                sec51["lf_cess"]  = cn(row[4]) if len(row) > 4 else None

    # ── Section 6.1 Payment ──
    pay_rows = []
    collecting_pay = False
    for tbl in all_tables:
        for row in tbl:
            if not row:
                continue
            row_text = " ".join(str(c or "") for c in row).lower()
            if "payment of tax" in row_text or "(a) other than reverse charge" in row_text:
                collecting_pay = True
            if collecting_pay:
                pay_rows.append(row)
    pay = parse_payment_table(pay_rows)

    # ── Page 3: Breakup of tax liability by period ──
    breakup = {}
    for tbl in all_tables:
        for row in tbl:
            if not row:
                continue
            label = str(row[0] or "").lower()
            if re.search(r'(january|february|march|april|may|june|july|august|september|october|november|december)\s+\d{4}', label):
                breakup = {
                    "period":   str(row[0] or "").strip(),
                    "igst":     cn(row[1]) if len(row) > 1 else None,
                    "cgst":     cn(row[2]) if len(row) > 2 else None,
                    "sgst":     cn(row[3]) if len(row) > 3 else None,
                    "cess":     cn(row[4]) if len(row) > 4 else None,
                }

    return {
        "meta":    meta,
        "3.1":     sec31,
        "3.1.1":   sec311,
        "3.2":     sec32,
        "4":       itc,
        "5":       sec5,
        "5.1":     sec51,
        "6.1":     pay,
        "breakup": breakup,
    }


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL WRITING
# ─────────────────────────────────────────────────────────────────────────────

MONTH_OFFSET = {
    4: 0, 5: 1, 6: 2, 7: 3, 8: 4, 9: 5,
    10: 6, 11: 7, 12: 8, 1: 9, 2: 10, 3: 11
}
MONTH_NUM = {
    "January": 1, "February": 2, "March": 3, "April": 4,
    "May": 5, "June": 6, "July": 7, "August": 8,
    "September": 9, "October": 10, "November": 11, "December": 12,
}


def w(ws, row, col, val):
    """Write val only if not None; never overwrite a formula cell."""
    if val is None:
        return
    existing = ws.cell(row=row, column=col).value
    if existing is not None and str(existing).startswith("="):
        return
    ws.cell(row=row, column=col).value = val


def write_sheet(ws, ex, month_num):
    off = MONTH_OFFSET[month_num]

    # ── 3.1 (rows 5–16, base 5) ──
    r = 5 + off
    a = ex["3.1"].get("a", {})
    b = ex["3.1"].get("b", {})
    c = ex["3.1"].get("c", {})
    d = ex["3.1"].get("d", {})
    e = ex["3.1"].get("e", {})

    # 3.1(a): B=taxable C=IGST D=CGST E=SGST(formula) F=Cess
    w(ws, r, 2, a.get("taxable"))
    w(ws, r, 3, a.get("igst"))
    w(ws, r, 4, a.get("cgst"))
    if a.get("sgst") is not None and a.get("cgst") is not None and abs(a["sgst"] - a["cgst"]) > 0.5:
        w(ws, r, 5, a["sgst"])
    w(ws, r, 6, a.get("cess"))

    # 3.1(b): G=taxable H=IGST I=Cess
    w(ws, r, 7, b.get("taxable"))
    w(ws, r, 8, b.get("igst"))
    w(ws, r, 9, b.get("cess"))

    # 3.1(c): J=value
    w(ws, r, 10, c.get("taxable"))

    # 3.1(d): K=taxable L=IGST M=CGST N=SGST(formula) O=Cess
    w(ws, r, 11, d.get("taxable"))
    w(ws, r, 12, d.get("igst"))
    w(ws, r, 13, d.get("cgst"))
    if d.get("sgst") is not None and d.get("cgst") is not None and abs(d["sgst"] - d["cgst"]) > 0.5:
        w(ws, r, 14, d["sgst"])
    w(ws, r, 15, d.get("cess"))

    # 3.1(e): P=value
    w(ws, r, 16, e.get("taxable"))

    # 3.2: Q=URD taxable R=URD IGST S=Comp taxable T=Comp IGST U=UIN taxable V=UIN IGST
    s32 = ex["3.2"]
    w(ws, r, 17, s32.get("urd",  {}).get("taxable"))
    w(ws, r, 18, s32.get("urd",  {}).get("igst"))
    w(ws, r, 19, s32.get("comp", {}).get("taxable"))
    w(ws, r, 20, s32.get("comp", {}).get("igst"))
    w(ws, r, 21, s32.get("uin",  {}).get("taxable"))
    w(ws, r, 22, s32.get("uin",  {}).get("igst"))

    # ── 4 ITC (rows 22–33, base 22) ──
    r4 = 22 + off
    itc = ex["4"]

    def iv(key, tax):
        return itc.get(key, {}).get(tax)

    # IGST: C=Available(4A5+4A4) D=RCM(4A3) E=Reversed(4B2+4B1) G=Ineligible(4D1)
    a5i = iv("4A5", "igst") or 0
    a4i = iv("4A4", "igst") or 0
    if a5i or a4i:
        ws.cell(r4, 3).value = f"={a5i}+{a4i}" if a4i else a5i

    w(ws, r4, 4, iv("4A3", "igst"))

    b1i = iv("4B1", "igst") or 0
    b2i = iv("4B2", "igst") or 0
    if b1i or b2i:
        ws.cell(r4, 5).value = f"={b2i}+{b1i}" if (b1i and b2i) else (b2i or b1i)
    elif iv("4B1", "igst") is not None or iv("4B2", "igst") is not None:
        ws.cell(r4, 5).value = 0

    w(ws, r4, 7, iv("4D1", "igst"))

    # CGST: K=Available(4A5) L=RCM(4A3) M=Reversed(4B1+4B2) O=Ineligible(4D1)
    w(ws, r4, 11, iv("4A5", "cgst"))
    w(ws, r4, 12, iv("4A3", "cgst"))

    b1c = iv("4B1", "cgst") or 0
    b2c = iv("4B2", "cgst") or 0
    if b1c or b2c:
        ws.cell(r4, 13).value = f"={b1c}+{b2c}" if (b1c and b2c) else (b1c or b2c)
    elif iv("4B1", "cgst") is not None or iv("4B2", "cgst") is not None:
        ws.cell(r4, 13).value = 0

    w(ws, r4, 15, iv("4D1", "cgst"))

    # SGST: cols S=19 T=20 U=21 W=23 mirror CGST via template formulas — do NOT write

    # CESS: AA=27 AB=28 AC=29
    w(ws, r4, 27, iv("4A5", "cess"))
    w(ws, r4, 28, iv("4A3", "cess"))
    b1ce = iv("4B1", "cess") or 0
    b2ce = iv("4B2", "cess") or 0
    if b1ce or b2ce:
        ws.cell(r4, 29).value = b1ce + b2ce

    # ── 5 (rows 40–51, base 40) ──
    r5 = 40 + off
    s5  = ex["5"]
    s51 = ex["5.1"]
    w(ws, r5, 2, s5.get("comp_inter"))
    w(ws, r5, 3, s5.get("comp_intra"))
    w(ws, r5, 4, s5.get("nongst_inter"))
    w(ws, r5, 5, s5.get("nongst_intra"))

    # 5.1 interest & late fee (same row block, cols F–M)
    w(ws, r5, 6,  s51.get("int_igst"))
    w(ws, r5, 7,  s51.get("int_cgst"))
    # Col H=8 is formula =+G; only write if sgst interest differs
    if s51.get("int_sgst") is not None and s51.get("int_cgst") is not None:
        if abs((s51["int_sgst"] or 0) - (s51["int_cgst"] or 0)) > 0.01:
            w(ws, r5, 8, s51["int_sgst"])
    w(ws, r5, 9,  s51.get("int_cess"))
    w(ws, r5, 10, s51.get("lf_igst"))
    w(ws, r5, 11, s51.get("lf_cgst"))
    # Col L=12 is formula =+K; only write if sgst late fee differs
    if s51.get("lf_sgst") is not None and s51.get("lf_cgst") is not None:
        if abs((s51["lf_sgst"] or 0) - (s51["lf_cgst"] or 0)) > 0.01:
            w(ws, r5, 12, s51["lf_sgst"])
    w(ws, r5, 13, s51.get("lf_cess"))

    # ── 6.1A IGST (rows 58–69, base 58) ──
    r6a = 58 + off
    igst_p = ex["6.1"].get("6A", {}).get("igst", {})
    # B=payable is formula; C=ITC_IGST D=ITC_CGST E=ITC_SGST F=cash G=interest
    w(ws, r6a, 3, igst_p.get("itc_igst"))
    w(ws, r6a, 4, igst_p.get("itc_cgst"))
    w(ws, r6a, 5, igst_p.get("itc_sgst"))
    w(ws, r6a, 6, igst_p.get("cash"))
    w(ws, r6a, 7, igst_p.get("interest"))

    # ── 6.1A Central + State Tax (rows 76–87, base 76) ──
    r6ct = 76 + off
    cgst_p = ex["6.1"].get("6A", {}).get("cgst", {})
    sgst_p = ex["6.1"].get("6A", {}).get("sgst", {})
    # Central: B=formula C=ITC_IGST D=ITC_CGST E=cash F=interest G=latefee
    w(ws, r6ct, 3, cgst_p.get("itc_igst"))
    w(ws, r6ct, 4, cgst_p.get("itc_cgst"))
    w(ws, r6ct, 5, cgst_p.get("cash"))
    w(ws, r6ct, 6, cgst_p.get("interest"))
    w(ws, r6ct, 7, cgst_p.get("latefee"))
    # State: I=formula J=ITC_IGST K=ITC_SGST L=cash M=interest
    w(ws, r6ct, 10, sgst_p.get("itc_igst"))
    w(ws, r6ct, 11, sgst_p.get("itc_sgst"))
    w(ws, r6ct, 12, sgst_p.get("cash"))
    w(ws, r6ct, 13, sgst_p.get("interest"))

    # ── 6.1B RCM (rows 94–105, base 94) ──
    r6b = 94 + off
    rcm_i = ex["6.1"].get("6B", {}).get("igst", {})
    rcm_c = ex["6.1"].get("6B", {}).get("cgst", {})
    # B=formula(IGST payable) C=IGST cash; E=formula(CGST) F=CGST cash; H=formula(SGST) I=formula =F
    w(ws, r6b, 3, rcm_i.get("cash"))
    w(ws, r6b, 6, rcm_c.get("cash"))


def update_turnover(wb, ex, month_num):
    if "Turnover" not in wb.sheetnames:
        return
    ws = wb["Turnover"]
    meta = ex["meta"]
    gstin = meta.get("gstin", "")
    year_str = meta.get("year", "") or ""
    period = meta.get("period", "")
    fy_start = int(year_str.split("-")[0]) if "-" in year_str else datetime.date.today().year
    year = fy_start if month_num >= 4 else fy_start + 1
    dt = datetime.datetime(year, month_num, 1)

    max_row = ws.max_row
    insert_row = max_row + 1
    for row in range(2, max_row + 1):
        if str(ws.cell(row, 2).value) == gstin and ws.cell(row, 3).value == dt:
            insert_row = row
            break

    s31  = ex["3.1"]
    s32  = ex["3.2"]
    rows_to_write = [
        ("3.1 (a) other than zero rated, nil rated and exempted",
         s31.get("a",{}).get("taxable",0), s31.get("a",{}).get("igst",0),
         s31.get("a",{}).get("cgst",0), s31.get("a",{}).get("sgst",0), s31.get("a",{}).get("cess",0)),
        ("3.1 (b) Zero Rated",
         s31.get("b",{}).get("taxable",0), s31.get("b",{}).get("igst",0), None, None, s31.get("b",{}).get("cess",0)),
        ("3.1 (c) Nil Rated, Exempted",
         s31.get("c",{}).get("taxable",0), None, None, None, None),
        ("3.1 (d) Liable to reverse charge",
         s31.get("d",{}).get("taxable",0), s31.get("d",{}).get("igst",0),
         s31.get("d",{}).get("cgst",0), s31.get("d",{}).get("sgst",0), s31.get("d",{}).get("cess",0)),
        ("3.1 (e) Non -GST Outward Supplies",
         s31.get("e",{}).get("taxable",0), None, None, None, None),
        ("URD Person",
         s32.get("urd",{}).get("taxable",0), s32.get("urd",{}).get("igst",0), None, None, None),
        ("Composition Person",
         s32.get("comp",{}).get("taxable",0), s32.get("comp",{}).get("igst",0), None, None, None),
        ("UIN",
         s32.get("uin",{}).get("taxable",0), s32.get("uin",{}).get("igst",0), None, None, None),
    ]
    for i, (sup, taxable, igst, cgst, sgst, cess) in enumerate(rows_to_write):
        r = insert_row + i
        ws.cell(r, 1).value = r - 1
        ws.cell(r, 2).value = gstin
        ws.cell(r, 3).value = dt
        ws.cell(r, 4).value = sup
        ws.cell(r, 5).value = taxable or 0
        ws.cell(r, 6).value = igst   or 0
        ws.cell(r, 7).value = cgst   or 0
        ws.cell(r, 8).value = sgst   or 0
        ws.cell(r, 9).value = cess   or 0


def process_pdfs(pdf_list, excel_bytes):
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    tmp.write(excel_bytes)
    tmp.flush()
    tmp.close()
    wb = load_workbook(tmp.name, keep_vba=False)
    results = []
    for fname, pdf_bytes in pdf_list:
        try:
            ex = extract_gstr3b(pdf_bytes)
            gstin = ex["meta"].get("gstin")
            period = ex["meta"].get("period", "")
            month_num = MONTH_NUM.get(period)
            if not gstin:
                results.append({"file": fname, "status": "❌ GSTIN not found",          "gstin": None,  "month": period, "ex": ex})
            elif not month_num:
                results.append({"file": fname, "status": "❌ Period not found",          "gstin": gstin, "month": period, "ex": ex})
            elif gstin not in wb.sheetnames:
                results.append({"file": fname, "status": f"⚠️ Sheet '{gstin}' missing", "gstin": gstin, "month": period, "ex": ex})
            else:
                write_sheet(wb[gstin], ex, month_num)
                update_turnover(wb, ex, month_num)
                results.append({"file": fname, "status": "✅ Mapped successfully",       "gstin": gstin, "month": period, "ex": ex})
        except Exception as e:
            import traceback
            results.append({"file": fname, "status": f"❌ Error: {e}", "gstin": None, "month": None, "ex": None, "tb": traceback.format_exc()})
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    os.unlink(tmp.name)
    return out.getvalue(), results


# ─────────────────────────────────────────────────────────────────────────────
# HELPERS FOR DISPLAY
# ─────────────────────────────────────────────────────────────────────────────

def fmt(v):
    if v is None:
        return "—"
    if isinstance(v, float):
        if v == 0:
            return "0"
        return f"{v:,.2f}"
    return str(v)


def section_table(rows, headers):
    """Render a compact HTML table for a section."""
    hdr_html = "".join(f"<th style='padding:4px 10px;text-align:right;background:#1e3a5f;color:white'>{h}</th>" for h in headers)
    body = ""
    for label, vals in rows:
        cells = f"<td style='padding:3px 8px;font-weight:500'>{label}</td>"
        for v in vals:
            color = "#c0392b" if isinstance(v, float) and v > 0 else "#2c3e50"
            cells += f"<td style='padding:3px 10px;text-align:right;color:{color}'>{fmt(v)}</td>"
        body += f"<tr style='border-bottom:1px solid #eee'>{cells}</tr>"
    return f"""<table style='width:100%;border-collapse:collapse;font-size:13px'>
    <thead><tr>{hdr_html}</tr></thead><tbody>{body}</tbody></table>"""


def show_full_extraction(ex):
    meta = ex["meta"]
    s31  = ex["3.1"]
    s311 = ex["3.1.1"]
    s32  = ex["3.2"]
    itc  = ex["4"]
    s5   = ex["5"]
    s51  = ex["5.1"]
    pay  = ex["6.1"]
    brk  = ex["breakup"]

    def iv(key, tax): return itc.get(key, {}).get(tax)
    def pv(sec, tax, fld): return pay.get(sec, {}).get(tax, {}).get(fld)

    # ── Meta ──
    st.markdown("**📋 Filing Details**")
    m1, m2, m3, m4 = st.columns(4)
    m1.metric("GSTIN",  meta.get("gstin", "—"))
    m2.metric("Period", f"{meta.get('period','')} {meta.get('year','')}")
    m3.metric("ARN",    meta.get("arn", "—"))
    m4.metric("Filed",  meta.get("date_arn", "—"))

    st.markdown("---")

    # ── 3.1 Outward Supplies ──
    st.markdown("**3.1 Outward & Inward Supplies**")
    rows_31 = [
        ("(a) Outward taxable (other than zero rated)", [
            s31.get("a",{}).get("taxable"), s31.get("a",{}).get("igst"),
            s31.get("a",{}).get("cgst"),   s31.get("a",{}).get("sgst"), s31.get("a",{}).get("cess")]),
        ("(b) Zero Rated", [
            s31.get("b",{}).get("taxable"), s31.get("b",{}).get("igst"),
            None, None, s31.get("b",{}).get("cess")]),
        ("(c) Nil Rated / Exempted", [s31.get("c",{}).get("taxable"), None, None, None, None]),
        ("(d) Inward supplies – Reverse Charge", [
            s31.get("d",{}).get("taxable"), s31.get("d",{}).get("igst"),
            s31.get("d",{}).get("cgst"),   s31.get("d",{}).get("sgst"), s31.get("d",{}).get("cess")]),
        ("(e) Non-GST Outward Supplies", [s31.get("e",{}).get("taxable"), None, None, None, None]),
    ]
    st.markdown(section_table(rows_31, ["", "Taxable Value ₹", "IGST ₹", "CGST ₹", "SGST ₹", "Cess ₹"]), unsafe_allow_html=True)

    # ── 3.1.1 ──
    st.markdown("**3.1.1 Section 9(5) e-Commerce**")
    rows_311 = [
        ("(i) ECO pays tax u/s 9(5)", [
            s311.get("i",{}).get("taxable"), s311.get("i",{}).get("igst"),
            s311.get("i",{}).get("cgst"),   s311.get("i",{}).get("sgst"), s311.get("i",{}).get("cess")]),
        ("(ii) Supplies through ECO", [s311.get("ii",{}).get("taxable"), None, None, None, None]),
    ]
    st.markdown(section_table(rows_311, ["", "Taxable ₹", "IGST ₹", "CGST ₹", "SGST ₹", "Cess ₹"]), unsafe_allow_html=True)

    # ── 3.2 ──
    st.markdown("**3.2 Inter-State Supplies**")
    rows_32 = [
        ("Supplies to Unregistered Persons", [s32.get("urd",{}).get("taxable"), s32.get("urd",{}).get("igst")]),
        ("Supplies to Composition Persons",  [s32.get("comp",{}).get("taxable"), s32.get("comp",{}).get("igst")]),
        ("Supplies to UIN Holders",          [s32.get("uin",{}).get("taxable"), s32.get("uin",{}).get("igst")]),
    ]
    st.markdown(section_table(rows_32, ["", "Taxable Value ₹", "IGST ₹"]), unsafe_allow_html=True)

    st.markdown("---")

    # ── 4 ITC ──
    st.markdown("**4. Eligible ITC**")
    itc_rows = [
        ("4A(1) Import of Goods",         [iv("4A1","igst"), iv("4A1","cgst"), iv("4A1","sgst"), iv("4A1","cess")]),
        ("4A(2) Import of Services",      [iv("4A2","igst"), iv("4A2","cgst"), iv("4A2","sgst"), iv("4A2","cess")]),
        ("4A(3) RCM inward (other)",      [iv("4A3","igst"), iv("4A3","cgst"), iv("4A3","sgst"), iv("4A3","cess")]),
        ("4A(4) Inward from ISD",         [iv("4A4","igst"), iv("4A4","cgst"), iv("4A4","sgst"), iv("4A4","cess")]),
        ("4A(5) All other ITC",           [iv("4A5","igst"), iv("4A5","cgst"), iv("4A5","sgst"), iv("4A5","cess")]),
        ("4B(1) Reversed – Rules 38/42/43",[iv("4B1","igst"), iv("4B1","cgst"), iv("4B1","sgst"), iv("4B1","cess")]),
        ("4B(2) Reversed – Others",       [iv("4B2","igst"), iv("4B2","cgst"), iv("4B2","sgst"), iv("4B2","cess")]),
        ("4C  Net ITC Available (A–B)",   [iv("4C","igst"),  iv("4C","cgst"),  iv("4C","sgst"),  iv("4C","cess")]),
        ("4D(1) ITC Reclaimed",           [iv("4D1","igst"), iv("4D1","cgst"), iv("4D1","sgst"), iv("4D1","cess")]),
        ("4D(2) Ineligible ITC s.16(4)",  [iv("4D2","igst"), iv("4D2","cgst"), iv("4D2","sgst"), iv("4D2","cess")]),
    ]
    st.markdown(section_table(itc_rows, ["", "IGST ₹", "CGST ₹", "SGST ₹", "Cess ₹"]), unsafe_allow_html=True)

    # ── Excel ITC mapping preview ──
    st.markdown("↳ *Excel writes: C = 4A5+4A4, D = 4A3, E = 4B2+4B1, G = 4D1  |  K = 4A5(CGST), L = 4A3(CGST), M = 4B1+4B2(CGST), O = 4D1(CGST)*")

    st.markdown("---")

    # ── 5 & 5.1 ──
    st.markdown("**5. Exempt / Non-GST Inward Supplies & 5.1 Interest / Late Fee**")
    c5a, c5b = st.columns(2)
    with c5a:
        rows_5 = [
            ("Composition / Exempt / Nil Rated", [s5.get("comp_inter"), s5.get("comp_intra")]),
            ("Non-GST Supply",                   [s5.get("nongst_inter"), s5.get("nongst_intra")]),
        ]
        st.markdown(section_table(rows_5, ["", "Inter-State ₹", "Intra-State ₹"]), unsafe_allow_html=True)
    with c5b:
        rows_51 = [
            ("Interest Paid", [s51.get("int_igst"), s51.get("int_cgst"), s51.get("int_sgst"), s51.get("int_cess")]),
            ("Late Fee",      [s51.get("lf_igst"),  s51.get("lf_cgst"),  s51.get("lf_sgst"),  s51.get("lf_cess")]),
        ]
        st.markdown(section_table(rows_51, ["", "IGST ₹", "CGST ₹", "SGST ₹", "Cess ₹"]), unsafe_allow_html=True)

    st.markdown("---")

    # ── 6.1 Payment ──
    st.markdown("**6.1A Payment of Tax – Other than Reverse Charge**")
    rows_6a = [
        ("IGST", [pv("6A","igst","payable"), pv("6A","igst","itc_igst"), pv("6A","igst","itc_cgst"),
                  pv("6A","igst","itc_sgst"), pv("6A","igst","cash"), pv("6A","igst","interest"), pv("6A","igst","latefee")]),
        ("CGST", [pv("6A","cgst","payable"), pv("6A","cgst","itc_igst"), pv("6A","cgst","itc_cgst"),
                  pv("6A","cgst","itc_sgst"), pv("6A","cgst","cash"), pv("6A","cgst","interest"), pv("6A","cgst","latefee")]),
        ("SGST", [pv("6A","sgst","payable"), pv("6A","sgst","itc_igst"), pv("6A","sgst","itc_cgst"),
                  pv("6A","sgst","itc_sgst"), pv("6A","sgst","cash"), pv("6A","sgst","interest"), pv("6A","sgst","latefee")]),
        ("CESS", [pv("6A","cess","payable"), pv("6A","cess","itc_igst"), pv("6A","cess","itc_cgst"),
                  pv("6A","cess","itc_sgst"), pv("6A","cess","cash"), pv("6A","cess","interest"), pv("6A","cess","latefee")]),
    ]
    st.markdown(section_table(rows_6a, ["Tax", "Payable ₹", "ITC-IGST ₹", "ITC-CGST ₹", "ITC-SGST ₹", "Cash ₹", "Interest ₹", "Late Fee ₹"]),
                unsafe_allow_html=True)

    st.markdown("**6.1B Payment of Tax – Reverse Charge**")
    rows_6b = [
        ("IGST", [pv("6B","igst","payable"), pv("6B","igst","cash")]),
        ("CGST", [pv("6B","cgst","payable"), pv("6B","cgst","cash")]),
        ("SGST", [pv("6B","sgst","payable"), pv("6B","sgst","cash")]),
        ("CESS", [pv("6B","cess","payable"), pv("6B","cess","cash")]),
    ]
    st.markdown(section_table(rows_6b, ["Tax", "Payable ₹", "Cash Paid ₹"]), unsafe_allow_html=True)

    # ── Breakup by period ──
    if brk:
        st.markdown("**Tax Liability Breakup (Period-wise)**")
        rows_brk = [(brk.get("period",""), [brk.get("igst"), brk.get("cgst"), brk.get("sgst"), brk.get("cess")])]
        st.markdown(section_table(rows_brk, ["Period", "IGST ₹", "CGST ₹", "SGST ₹", "Cess ₹"]), unsafe_allow_html=True)


# ─────────────────────────────────────────────────────────────────────────────
# STREAMLIT UI
# ─────────────────────────────────────────────────────────────────────────────

st.title("📊 GSTR-3B → Excel Auto-Mapper")
st.markdown("Upload GSTR-3B PDFs and your GST Monthly Analysis Excel. All fields are extracted and previewed before mapping.")

col1, col2 = st.columns(2)

with col1:
    st.subheader("1. Upload Excel Template")
    excel_file = st.file_uploader("GST Monthly Analysis (.xlsx)", type=["xlsx"])
    if excel_file:
        wb_check = load_workbook(io.BytesIO(excel_file.read()), data_only=False)
        gstin_sheets = [s for s in wb_check.sheetnames if re.match(r'^\d{2}[A-Z]{5}\d{4}', s)]
        st.success(f"✅ **{excel_file.name}** — {len(gstin_sheets)} GSTIN sheets")
        excel_file.seek(0)

with col2:
    st.subheader("2. Upload GSTR-3B PDFs")
    pdf_files = st.file_uploader("GSTR-3B PDFs (one or many)", type=["pdf"], accept_multiple_files=True)
    if pdf_files:
        st.success(f"✅ {len(pdf_files)} PDF(s) loaded")
        for f in pdf_files:
            st.caption(f"• {f.name}")

st.divider()

if excel_file and pdf_files:
    st.subheader("3. Extracted Data — Full Preview")
    st.caption("Every field extracted from each PDF is shown below. Scroll through each section before mapping.")

    with st.spinner("Extracting all data from PDFs..."):
        previews = []
        for pf in pdf_files:
            try:
                ex = extract_gstr3b(pf.read())
                previews.append({"file": pf.name, "ex": ex, "err": None})
            except Exception as e:
                previews.append({"file": pf.name, "ex": None, "err": str(e)})
            pf.seek(0)

    for item in previews:
        gstin = item["ex"]["meta"].get("gstin", "?") if item["ex"] else "?"
        period = item["ex"]["meta"].get("period", "") if item["ex"] else ""
        with st.expander(f"📄 **{item['file']}**  |  GSTIN: {gstin}  |  {period}", expanded=True):
            if item["err"]:
                st.error(item["err"])
            else:
                show_full_extraction(item["ex"])

    st.divider()
    st.subheader("4. Map to Excel & Download")

    if st.button("🚀 Map All & Download", type="primary", use_container_width=True):
        with st.spinner("Writing to Excel..."):
            pdf_data = [(pf.name, pf.read()) for pf in pdf_files]
            excel_bytes = excel_file.read()
            out_bytes, results = process_pdfs(pdf_data, excel_bytes)

        for res in results:
            s = res["status"]
            label = f"**{res['file']}** ({res.get('gstin','?')} — {res.get('month','?')}): {s}"
            if "✅" in s:   st.success(label)
            elif "⚠️" in s: st.warning(label)
            else:
                st.error(label)
                if "tb" in res: st.code(res["tb"])

        fname_out = f"GST_Updated_{datetime.datetime.now():%Y%m%d_%H%M%S}.xlsx"
        st.download_button(
            "⬇️ Download Updated Excel",
            data=out_bytes,
            file_name=fname_out,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary",
        )
        st.caption("All existing formulas preserved. Only hardcoded data cells are updated.")

elif not excel_file and not pdf_files:
    st.info("Upload your Excel template and GSTR-3B PDFs above to begin.")
elif not excel_file:
    st.warning("Please upload the Excel template.")
else:
    st.warning("Please upload at least one GSTR-3B PDF.")